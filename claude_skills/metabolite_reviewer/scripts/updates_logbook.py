"""Convert a curated review sheet into rows of ``updates_to_model.xlsx``.

The logbook is the permanent record of every change made to the model, across
every review round and every agent that touches it. Three consequences shape
this module:

1.  **It is appended to, never rewritten.** Other agents (reaction review,
    gap-filling) write their own rows into the same file, and rounds that have
    already been applied must stay readable exactly as they were applied. The
    writer here reads the existing sheets, adds rows, and writes the file
    back whole — it never starts from a blank workbook, and it never edits a
    row it did not just create.
2.  **Only accepted rows are written.** A finding is actioned when
    ``fix_accepted`` is TRUE. Anything else — FALSE, blank, a comment, a
    curator who left the row alone — is skipped. A review the curator has not
    finished therefore produces a short logbook, not a wrong one.
3.  **The decision verb, not the finding, decides what is written.** The
    review sheet's ``proposed_fix`` is prose for a human. The verb in
    ``curator_decision`` is what maps onto a sheet and a set of columns. A row
    accepted without a recognised verb is reported back to the caller rather
    than guessed at.

Round numbering is taken from the existing file: the new rows get
``max(round) + 1`` unless a round is given explicitly, so the logbook stays
ordered even if rounds are applied out of sequence.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import polars as pl

try:                                  # normal use: alongside workbook.py
    from workbook import DECISION_VERBS
except ImportError:                   # imported as a package
    from .workbook import DECISION_VERBS

#: Sheet -> column order. Must match the template the curator supplied; a
#: sheet is only ever appended to, so the column order is fixed by the file.
SHEET_COLUMNS = {
    "new_metabolites": ["round", "id", "name", "formula", "charge",
                        "metanetx.chemical", "metacyc.compound", "inchikey",
                        "rationale"],
    "new_reactions": ["round", "id", "name", "reaction", "lower_bound",
                      "upper_bound", "gpr", "metanetx.reaction", "ec-code",
                      "metacyc.reaction", "group"],
    "changes": ["round", "entity_type", "change_type", "id", "replaced_by",
                "rationale"],
    "annotation": ["round", "entity_type", "id", "field", "value", "method",
                   "rationale"],
}

#: How this module identifies itself in the `method` column.
METHOD = "metabolite_reviewer"


def _read_sheets(path: Path) -> dict[str, pl.DataFrame]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            out[name] = pl.DataFrame()
            continue
        # The template carries trailing empty header cells, and openpyxl
        # returns short tuples for rows that end in blanks — so pad every row
        # to the header width before indexing into it.
        header = [str(h) for h in rows[0] if h is not None]
        width = len(header)
        body = []
        for r in rows[1:]:
            padded = tuple(r[:width]) + (None,) * max(0, width - len(r))
            if any(v is not None for v in padded):
                body.append(padded)
        out[name] = pl.DataFrame(
            {h: [str(r[i]) if r[i] is not None else None for r in body]
             for i, h in enumerate(header)},
            schema={h: pl.String for h in header}) if body else pl.DataFrame(
            {h: pl.Series([], dtype=pl.String) for h in header})
    wb.close()
    return out


def _next_round(sheets: dict[str, pl.DataFrame]) -> int:
    seen = []
    for frame in sheets.values():
        if "round" in frame.columns and frame.height:
            vals = (frame["round"].cast(pl.String).str.extract(r"(\d+)")
                    .cast(pl.Int64, strict=False).drop_nulls().to_list())
            seen.extend(vals)
    return (max(seen) + 1) if seen else 1


def rows_from_review(review: pl.DataFrame, mets: pl.DataFrame,
                     round_no: int, id_col: str = "id"
                     ) -> tuple[dict[str, list[dict]], list[dict]]:
    """Split accepted review rows into logbook rows. Returns (rows, rejected).

    ``rejected`` holds accepted rows this module could not translate — an
    unrecognised verb, or a verb whose ``curator_note`` argument is missing.
    They are returned rather than dropped: a silently skipped acceptance is a
    change the curator believes was made and was not.
    """
    out: dict[str, list[dict]] = {k: [] for k in SHEET_COLUMNS}
    bad: list[dict] = []
    lookup = {r[id_col]: r for r in mets.iter_rows(named=True)} \
        if id_col in mets.columns else {}

    accepted = review.filter(
        pl.col("fix_accepted").cast(pl.String).str.strip_chars()
          .str.to_uppercase() == "TRUE")

    for row in accepted.iter_rows(named=True):
        mid = row[id_col]
        verb = (row.get("curator_decision") or "").strip()
        note = (row.get("curator_note") or "").strip()
        why = f"round {round_no} metabolite review: {row.get('error_tag')}" \
              + (f" — {note}" if note else "")

        if verb not in DECISION_VERBS:
            bad.append({**row, "_reason": f"unrecognised decision {verb!r}"})
            continue
        if verb in ("reject", "defer_to_reaction_review"):
            continue
        # Some findings are about the sheet rather than about a species
        # (schema anomalies carry a placeholder id). A logbook row keyed on
        # one of those would instruct the updater to change a metabolite that
        # does not exist.
        if lookup and mid not in lookup:
            bad.append({**row, "_reason": f"{mid!r} is not a metabolite in "
                                          "the model; not actionable here"})
            continue

        if verb in ("merge_into", "replace_with", "delete", "set_formula",
                    "set_charge", "set_name", "set_annotation"):
            # An explicit verb takes its argument from curator_note; if the
            # curator left it blank but the check already proposed the same
            # verb, its argument stands (the curator agreed with the fix and
            # only restated it explicitly).
            arg = note or ((row.get("fix_arg") or "").strip()
                           if (row.get("fix_verb") or "").strip() == verb
                           else "")
            translated, err = _apply_verb(verb, arg, mid, round_no, why)
            if err:
                bad.append({**row, "_reason": f"{err} in curator_note"})
                continue
            for sheet, entry in translated:
                out[sheet].append(entry)
        elif verb == "keep_both":
            out["annotation"].append({
                "round": round_no, "entity_type": "metabolite", "id": mid,
                "field": "curation_note", "value": "distinct entity; "
                "reviewed and kept", "method": METHOD, "rationale": why})
        elif verb == "accept":
            # `accept` means "apply the fix as proposed". The prose in
            # proposed_fix is for the human; the machine-readable form is the
            # check's own fix_verb/fix_arg. Parsing the prose was tried and
            # abandoned — it silently mistranslated merges — so a finding
            # whose check could not express its fix structurally is reported
            # back instead, and the curator names a verb.
            sub_verb = (row.get("fix_verb") or "").strip()
            sub_arg = (row.get("fix_arg") or "").strip()
            if not sub_verb:
                bad.append({**row, "_reason": "accept: this check does not "
                            "propose a machine-applicable fix; choose an "
                            "explicit verb (merge_into / set_* / delete)"})
                continue
            translated, err = _apply_verb(sub_verb, sub_arg, mid, round_no, why)
            if err:
                bad.append({**row, "_reason": f"accept: {err}"})
                continue
            for sheet, entry in translated:
                out[sheet].append(entry)

    # One metabolite can be flagged by several checks — a species can duplicate
    # its neighbour by formula *and* stand in for it in two different anchored
    # reactions. Those are three findings but one change, and writing it three
    # times would make the logbook read as three separate edits to the model.
    # Identity here is the change itself; the rationales are merged onto it.
    for sheet, rows in out.items():
        keyed: dict[tuple, dict] = {}
        for r in rows:
            k = tuple(str(r.get(c)) for c in SHEET_COLUMNS[sheet]
                      if c != "rationale")
            if k in keyed:
                prev = keyed[k]
                extra = str(r.get("rationale") or "")
                if extra and extra not in str(prev.get("rationale") or ""):
                    prev["rationale"] = f"{prev['rationale']}; also {extra}"
            else:
                keyed[k] = r
        out[sheet] = list(keyed.values())
    return out, bad


def _apply_verb(verb: str, arg: str, mid: str, round_no: int,
                why: str) -> tuple[list[tuple[str, dict]], str | None]:
    """Translate a verb + argument into logbook rows. (rows, error)."""
    if verb in ("merge_into", "replace_with"):
        if not arg:
            return [], f"{verb} needs a target id"
        return [("changes", {"round": round_no, "entity_type": "metabolite",
                             "change_type": "deletion", "id": mid,
                             "replaced_by": arg, "rationale": why})], None
    if verb == "delete":
        return [("changes", {"round": round_no, "entity_type": "metabolite",
                             "change_type": "deletion", "id": mid,
                             "replaced_by": None, "rationale": why})], None
    if verb in ("set_formula", "set_charge", "set_name"):
        if not arg:
            return [], f"{verb} needs a value"
        return [("annotation", {"round": round_no, "entity_type": "metabolite",
                                "id": mid, "field": verb.removeprefix("set_"),
                                "value": arg, "method": METHOD,
                                "rationale": why})], None
    if verb == "set_annotation":
        if "=" not in arg:
            return [], "set_annotation needs field=value"
        field, _, value = arg.partition("=")
        return [("annotation", {"round": round_no, "entity_type": "metabolite",
                                "id": mid, "field": field.strip(),
                                "value": value.strip(), "method": METHOD,
                                "rationale": why})], None
    return [], f"cannot apply {verb!r}"


def update_logbook(path, review: pl.DataFrame, mets: pl.DataFrame,
                   round_no: int | None = None, id_col: str = "id",
                   backup: bool = True) -> dict:
    """Append accepted decisions to ``updates_to_model.xlsx``.

    Returns a report: rows written per sheet, the round number used, and any
    accepted rows that could not be translated.

    The file must already exist — per the curator's instruction this logbook
    is shared across agents and rounds, so creating one silently would risk
    starting a second, competing history. Ask for the path instead.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"{path} not found. The updates logbook is the shared record for "
            "every review round; ask the curator for the existing file rather "
            "than starting a new one.")

    sheets = _read_sheets(path)
    round_no = round_no if round_no is not None else _next_round(sheets)
    new_rows, rejected = rows_from_review(review, mets, round_no, id_col)

    if backup:
        shutil.copy(path, path.with_suffix(".bak.xlsx"))

    # Re-running a round must not double its history. The logbook is the
    # record of what was done to the model, so a change already recorded for
    # this round is not written again — that makes the whole append idempotent
    # and safe to re-run after a crash or a partial curation pass.
    written, skipped = {}, 0
    for name, rows in new_rows.items():
        existing_keys = set()
        prior = sheets.get(name)
        cols = SHEET_COLUMNS[name]
        if prior is not None and prior.height:
            key_cols = [c for c in cols if c != "rationale"
                        and c in prior.columns]
            for r in prior.select(key_cols).iter_rows(named=True):
                existing_keys.add(tuple(
                    str(r[c]) if r[c] is not None else "None"
                    for c in key_cols))
            kept = []
            for r in rows:
                k = tuple(str(r.get(c)) for c in key_cols)
                if k in existing_keys:
                    skipped += 1
                else:
                    kept.append(r)
            rows = kept
        if not rows:
            continue
        add = pl.DataFrame([{c: (str(r[c]) if r.get(c) is not None else None)
                             for c in cols} for r in rows],
                           schema={c: pl.String for c in cols})
        existing = sheets.get(name)
        if existing is not None and existing.height:
            keep = [c for c in cols if c in existing.columns]
            existing = existing.select(keep).with_columns(
                [pl.lit(None, dtype=pl.String).alias(c)
                 for c in cols if c not in keep]).select(cols)
            sheets[name] = pl.concat([existing, add])
        else:
            sheets[name] = add
        written[name] = len(rows)

    import xlsxwriter as xw
    wb = xw.Workbook(str(path))
    for name in list(SHEET_COLUMNS) + [n for n in sheets
                                       if n not in SHEET_COLUMNS]:
        frame = sheets.get(name)
        if frame is None:
            frame = pl.DataFrame({c: pl.Series([], dtype=pl.String)
                                  for c in SHEET_COLUMNS.get(name, ["id"])})
        frame.write_excel(workbook=wb, worksheet=name, autofit=True,
                          header_format={"bold": True, "bg_color": "#DDDDDD"},
                          freeze_panes=(1, 0))
    wb.close()

    return {"round": round_no, "written": written,
            "already_present": skipped,
            "n_accepted": int(review.filter(
                pl.col("fix_accepted").cast(pl.String).str.strip_chars()
                  .str.to_uppercase() == "TRUE").height),
            "rejected": rejected}
